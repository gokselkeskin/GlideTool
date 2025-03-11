import matplotlib
import scipy as sc
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import numpy as np
from scipy.optimize import minimize
import tkinter as tk
from tkinter import simpledialog, messagebox
import csv
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
import sys
import io


if sys.stdout:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
else:
    sys.stdout = open(os.devnull, "w")

if sys.stderr:
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
else:
    sys.stderr = open(os.devnull, "w")

# Get the user's Documents folder
user_documents = os.path.join(os.path.expanduser("~"), "Documents")

subfolder = "GlideTool"
folder_path = os.path.join(user_documents, subfolder)

if not os.path.exists(folder_path):
    os.makedirs(folder_path)


def f_linear(x,a,b):
    return (a * x) + b


def beta_opt(Vt, m, g, rho, Sw, B, k, Cdpro, delta=1):
    try:
        value = np.clip(((8 * k * m ** 2 * g ** 2)
                         / (delta * np.pi * rho ** 2 * B ** 2 * Cdpro * Sw * Vt ** 4)
                         ) ** (1 / 3),
                        0, 1, )
    except RuntimeWarning:
        value = np.inf

    return value

def beta_hat(Vt, m, g, rho, Sw, B, k, Cdpro, delta=1):
    try:
        value = ((8 * k * m ** 2 * g ** 2)
                 / (delta * np.pi * rho ** 2 * B ** 2 * Cdpro * Sw * Vt ** 4)) ** (1 / 3)
    except RuntimeWarning:
        value = np.inf

    return value

def Di(Vt, k, m, g, B, rho, Sw, Cdpro, delta):
    beta = beta_opt(Vt, m, g, rho, Sw, B, k, Cdpro, delta)
    value = ((2 * k * m ** 2 * g ** 2)
             / (Vt ** 2 * np.pi * (beta * B) ** 2 * rho)
             )
    return value


def Db(Vt, rho, Sb, Cdb):
    value = rho * Vt ** 2 * Sb * Cdb / 2
    return value


def Dpro(Vt, rho, Sw, Cdpro, m, g, B, k, delta):
    beta = beta_opt(Vt, m, g, rho, Sw, B, k, Cdpro, delta)
    epsilon = 1 - delta * (1 - beta)
    value = rho * Vt ** 2 * epsilon * Sw * Cdpro / 2
    return value


def Vz_pennycuick(Vt, m, g, rho, Sb, Sw, B, Cdpro=0.014, Cdb=0.1, delta=1, e=1):
    AR = B ** 2 / Sw

    k= induced_drag_factor
    beta = np.clip(beta_opt(Vt, m, g, rho, Sw, B, k, Cdpro, delta), 0, 1, )


    epsilon = 1 - delta * (1 - beta)

    value = ((Di(Vt, k, m, g, B, rho, Sw, Cdpro=Cdpro, delta=delta)
              + Db(Vt=Vt, rho=rho, Sb=Sb, Cdb=Cdb)
              + Dpro(Vt=Vt, rho=rho, Sw=Sw, Cdpro=Cdpro, m=m, g=g, B=B, k=k, delta=delta)) * Vt
             / (m * g))
    return value


def input_bird_data():
    bird_data = {}

    root = tk.Tk()
    root.title("Bird Data Input")
    root.geometry("800x550")

    entry_width = 50
    pad_x, pad_y = 10, 4

    # Create a dictionary
    entries = {}

    note_label3 = tk.Label(root,
                           text="This is the GlideTool created as described in 'Data-driven optimization of Pennycuick’s Flight Tool for improved polar curves of thermal soaring bird species.'",
                           wraplength=800, justify="left")
    note_label3.grid(row=13, column=0, columnspan=2, padx=pad_x, pady=pad_y, sticky=tk.W)

    note_label = tk.Label(root,
                          text="Please enter all the bird details in the fields above. 'Falco naumanni (Fn)' is provided as an example in one of our datasets. All birds in the dataset are also available in the dropdown menu.",
                          wraplength=800, justify="left")
    note_label.grid(row=14, column=0, columnspan=2, padx=pad_x, pady=pad_y, sticky=tk.W)

    note_label2 = tk.Label(root,
                          text=f"Your results will be saved to: {folder_path}",
                          wraplength=500, justify="left")
    note_label2.grid(row=15, column=0, columnspan=2, padx=pad_x, pady=pad_y, sticky=tk.W)



    predefined_birds = {
        "Falco naumanni (Fn)": {
            "bird_name": "Fn",
            "mass": "0.13",
            "body_surface": "0.002089",
            "wing_area": "0.062",
            "wingspan": "0.68",
            "horizontal_velocities": "5., 6., 7., 8., 9., 10., 11., 12., 13., 14., 15., 16., 17., 18., 19.",
            "vertical_velocities": "-0.40511774, -0.66352524, -0.63802253, -0.730276, -0.90073146, -1.12255318, -1.42933517, -1.81373964, -2.12802013, -2.47379687, -2.8627934, -3.54308685, -3.47981722, -3.58184766, -4.24346973",
            "vertical_errors": "0.02050686, 0.02928516, 0.02399187, 0.01810837, 0.01634837, 0.01670776, 0.01673685, 0.01913213, 0.02081005, 0.0244458, 0.03374285, 0.05714717, 0.08958036, 0.12822351, 0.13210858"
        },
        "Falco peregrinus (Fp)": {
            "bird_name": "Fp",
            "mass": "0.815",
            "body_surface": "0.007123",
            "wing_area": "0.12785",
            "wingspan": "1.04",
            "horizontal_velocities": "7.,  8.,  9., 10., 11., 12., 13., 14., 15., 16., 17., 18., 19.",
            "vertical_velocities": "-1.31676286, -1.23181437, -1.2794416 , -1.04062891, -1.0851267 ,-1.23557091, -1.37566544, -1.46694618, -1.51697174, -1.91683229,-2.16842994, -2.52912355, -2.8078217",
            "vertical_errors": "0.18085605, 0.09849189, 0.08734052, 0.07387672, 0.04913297,0.04002263, 0.03646723, 0.03455059, 0.03443361, 0.04105675,0.04356088, 0.05223451, 0.06250747"
        },
        "Aquila rapax (Ar)": {
            "bird_name": "Ar",
            "mass": "2.6",
            "body_surface": "0.01536",
            "wing_area": "0.5",
            "wingspan": "1.82",
            "horizontal_velocities": "8.,  9., 10., 11., 12., 13., 14., 15., 16., 17., 18., 19.",
            "vertical_velocities": "-0.77323757, -0.97679889, -1.03358537, -1.26375504, -1.43577333,-1.72671286, -2.11739609, -2.3944525 , -2.74356857, -3.04383433,-3.35187796, -3.27693643",
            "vertical_errors": "0.08425824, 0.05289849, 0.04289992, 0.0419337 , 0.0385476 ,0.04545141, 0.05223856, 0.06162638, 0.06552755, 0.09108939,0.13286914, 0.19473848"
        },
        "Aquila nipalensis (An)": {
            "bird_name": "An",
            "mass": "2.35",
            "body_surface": "0.01436",
            "wing_area": "0.54",
            "wingspan": "1.9",
            "horizontal_velocities": "5., 6., 7., 8., 9., 10., 11., 12., 13., 14., 15., 16., 17., 18., 19.",
            "vertical_velocities": "-0.43033187, -0.5392118, -0.62207852, -0.58565906, -0.62060839, -0.74807133, -0.88294852, -1.11862157, -1.38993983, -1.60650609, -1.87736075, -2.16980881, -2.44521326, -2.63897619, -2.53538794",
            "vertical_errors": "0.02131003, 0.01669836, 0.01409964, 0.00895406, 0.00791837, 0.00689004, 0.00657965, 0.0074163, 0.00930995, 0.01149468, 0.01532249, 0.01923124, 0.025801, 0.03641581, 0.05560339"
        },
        "Aquila verreauxii (Av)": {
            "bird_name": "Av",
            "mass": "3.3",
            "body_surface": "0.01801",
            "wing_area": "0.51",
            "wingspan": "2",
            "horizontal_velocities": "9., 10., 11., 12., 13., 14., 15., 16., 17., 18.",
            "vertical_velocities": "-1.51312336, -1.32579748, -1.5593121, -1.43500331, -1.34786955, -1.62101318, -1.71726104, -2.15343736, -1.80968474, -2.15167593",
            "vertical_errors": "0.31095606, 0.19516012, 0.19119541, 0.15393517, 0.11043504, 0.0940244, 0.08685104, 0.0701672, 0.07648489, 0.08421064"
        },
        "Haliaeetus leucocephalus (Hl)": {
            "bird_name": "Hl",
            "mass": "3.1",
            "body_surface": "0.01727",
            "wing_area": "0.57",
            "wingspan": "2",
            "horizontal_velocities": "7., 8., 9., 10., 11., 12., 13., 14., 15., 16., 17., 18.",
            "vertical_velocities": "-1.9088342, -1.45090718, -1.42033333, -1.41831052, -1.53360514, -1.74291959, -1.92674152, -1.98433498, -2.30916695, -2.52710672, -2.67241881, -2.98217356",
            "vertical_errors": "0.16299378, 0.11544433, 0.08977667, 0.06747671, 0.05717423, 0.05258846, 0.04502741, 0.04208638, 0.04631182, 0.05215592, 0.05144963, 0.06212867"
        },
        "Gyps fulvus (Gf1)": {
            "bird_name": "Gf1",
            "mass": "7.7",
            "body_surface": "0.03166",
            "wing_area": "0.95",
            "wingspan": "2.56",
            "horizontal_velocities": "6., 7., 8., 9., 10., 11., 12., 13., 14., 15., 16., 17., 18., 19.",
            "vertical_velocities": "-4.46301691, -3.19145604, -2.27305599, -1.69067704, -1.55617669, -1.43166432, -1.41160416, -1.53032998, -1.72281932, -1.87687444, -2.04586206, -2.19869584, -2.39468402, -2.69739784",
            "vertical_errors": "0.29814289, 0.17467505, 0.08989459, 0.05118043, 0.03029308, 0.01990733, 0.01443676, 0.01350607, 0.01297754, 0.01294294, 0.01304375, 0.01338602, 0.01586135, 0.01925413"
        },
        "Gyps fulvus (Gf2)": {
            "bird_name": "Gf2",
            "mass": "8.5",
            "body_surface": "0.03381",
            "wing_area": "0.95",
            "wingspan": "2.7",
            "horizontal_velocities": "7., 8., 9., 10., 11., 12., 13., 14., 15., 16., 17., 18., 19.",
            "vertical_velocities": "-1.57865009, -1.48425895, -1.38948447, -1.32873572, -1.31640848, -1.39704216, -1.4920292, -1.62165779, -1.74228166, -1.88942302, -2.00778894, -2.14385164, -2.26872903",
            "vertical_errors": "0.03276291, 0.02249218, 0.01476003, 0.00961578, 0.00708564, 0.00558237, 0.00469789, 0.00420678, 0.00389111, 0.003787, 0.00386621, 0.00420599, 0.00499275"
        },
        "Gyps himalayensis (Gh)": {
            "bird_name": "Gh",
            "mass": "8",
            "body_surface": "0.03247",
            "wing_area": "1.25",
            "wingspan": "2.8",
            "horizontal_velocities": "6., 7., 8., 9., 10., 11., 12., 13., 14., 15., 16., 17., 18., 19.",
            "vertical_velocities": "-3.30137262, -1.93561767, -1.94652412, -1.58686272, -1.41474445, -1.37137819, -1.49091112, -1.66854518, -1.761876, -1.96227801, -2.09080322, -2.18425048, -2.44890737, -2.61074866",
            "vertical_errors": "0.18325344, 0.0972943, 0.05410993, 0.03147664, 0.01966184, 0.01511639, 0.01240439, 0.01137316, 0.0104692, 0.01033173, 0.01059433, 0.01091893, 0.01249874, 0.01528492"
        },
        "Gyps rueppellii (Gr)": {
            "bird_name": "Gr",
            "mass": "5.6",
            "body_surface": "0.02561",
            "wing_area": "0.75",
            "wingspan": "2.2",
            "horizontal_velocities": "8., 9., 10., 11., 12., 13., 14., 15., 16., 17., 18.",
            "vertical_velocities": "-3.08411134, -1.8432253, -1.65151927, -1.72873715, -1.70089437, -1.88405321, -2.08846381, -2.21972699, -2.31803006, -2.45402351, -2.88234534",
            "vertical_errors": "0.26336011, 0.10796004, 0.06677896, 0.04972926, 0.02958507, 0.02895477, 0.03891346, 0.03740936, 0.0423801, 0.03809539, 0.04467858"
        },
        "Vultur gryphus (Vg)": {
            "bird_name": "Vg",
            "mass": "10",
            "body_surface": "0.03768",
            "wing_area": "1.099",
            "wingspan": "2.9",
            "horizontal_velocities": "9., 10., 11., 12., 13., 14., 15., 16., 17., 18.",
            "vertical_velocities": "-1.92861235, -1.52396454, -1.28788243, -1.00051013, -1.17357668, -1.47610497, -1.76936261, -1.85681083, -2.39867139, -2.80159964",
            "vertical_errors": "0.21482829, 0.09060667, 0.07205334, 0.04918992, 0.04196249, 0.04684111, 0.06852593, 0.07148479, 0.10071945, 0.10126608"
        },

        "Ciconia ciconia (Cc)": {
            "bird_name": "Cc",
            "mass": "3.5",
            "body_surface": "0.01873",
            "wing_area": "0.6",
            "wingspan": "2",
            "horizontal_velocities": "6., 7., 8., 9., 10., 11., 12., 13., 14., 15., 16., 17., 18.",
            "vertical_velocities": "-0.88370495, -0.96413152, -0.94731884, -0.94882002, -0.99076135, -1.11037044, -1.15184995, -1.20193306, -1.26526952, -1.33347239, -1.38693865, -1.52849487, -1.69289299",
            "vertical_errors": "0.0175761, 0.01503057, 0.0102909, 0.0070037, 0.0056166, 0.00502337, 0.00431528, 0.00424222, 0.00422315, 0.00449917, 0.00495638, 0.00639321, 0.0089695"
        },
        "Geronticus eremita (Ge)": {
            "bird_name": "Ge",
            "mass": "1.2",
            "body_surface": "0.00918",
            "wing_area": "0.23",
            "wingspan": "1.3",
            "horizontal_velocities": "5., 6., 7., 8., 9., 10., 11., 12., 13., 14., 15., 16., 17., 18.",
            "vertical_velocities": "-0.48496027, -0.49387826, -0.46641524, -0.46280666, -0.43688456, -0.52255435, -0.61456078, -0.7360827, -0.89799537, -1.03922923, -1.16363944, -1.42437075, -1.59229132, -1.48132091",
            "vertical_errors": "0.04016607, 0.020232, 0.01895447, 0.0153938, 0.009226, 0.00719243, 0.0065559, 0.0061327, 0.00638075, 0.00687022, 0.00717367, 0.0089231, 0.01074643, 0.01617643"
        }

    }

    bird_options = list(predefined_birds.keys()) + ["Custom Entry"]
    selected_bird = tk.StringVar(root)
    selected_bird.set("Falco naumanni (Fn)")  # Default selection

    def update_fields(*args):
        bird = selected_bird.get()
        if bird in predefined_birds:
            for key, value in predefined_birds[bird].items():
                entries[key].delete(0, tk.END)
                entries[key].insert(0, value)
        else:
            for entry in entries.values():
                entry.delete(0, tk.END)

    tk.Label(root, text="Select Bird Example:").grid(row=0, column=0, padx=pad_x, pady=pad_y, sticky=tk.W)
    dropdown = tk.OptionMenu(root, selected_bird, *bird_options, command=update_fields)
    dropdown.grid(row=0, column=1, padx=pad_x, pady=pad_y, sticky=tk.W)

    #input
    entries = {}
    labels = [
        ("bird_name", "Enter the bird's name:"),
        ("mass", "Enter the mass of the bird (kg):"),
        ("body_surface", "Enter the body surface area (m²):"),
        ("wing_area", "Enter the wing area (m²):"),
        ("wingspan", "Enter the wingspan (m):"),
        ("horizontal_velocities", "Enter horizontal velocities (comma-separated):"),
        ("vertical_velocities", "Enter vertical velocities (comma-separated):"),
        ("vertical_errors", "Enter vertical velocity errors (comma-separated, optional):")
    ]

    for i, (key, label) in enumerate(labels):
        tk.Label(root, text=label).grid(row=i + 1, column=0, padx=pad_x, pady=pad_y, sticky=tk.W)
        entry = tk.Entry(root, width=entry_width)
        entry.grid(row=i + 1, column=1, padx=pad_x, pady=pad_y, sticky=tk.EW)
        entries[key] = entry  # Store in dictionary

    update_fields()

    # Checkboxes for saving options
    save_fig_var = tk.IntVar()
    save_csv_var = tk.IntVar()
    save_excel_var = tk.IntVar()

    save_fig_check = tk.Checkbutton(root, text="Save Figure", variable=save_fig_var)
    save_fig_check.grid(row=len(labels) + 2, column=1, padx=pad_x, pady=pad_y, sticky=tk.W)

    save_csv_check = tk.Checkbutton(root, text="Save CSV", variable=save_csv_var)
    save_csv_check.grid(row=len(labels) + 3, column=1, padx=pad_x, pady=pad_y, sticky=tk.W)

    save_excel_check = tk.Checkbutton(root, text="Save Excel", variable=save_excel_var)
    save_excel_check.grid(row=len(labels) + 4, column=1, padx=pad_x, pady=pad_y, sticky=tk.W)

    def submit_data():
        try:
            bird_name = entries["bird_name"].get()
            m = float(entries["mass"].get())
            Sb = float(entries["body_surface"].get())
            Sw = float(entries["wing_area"].get())
            B = float(entries["wingspan"].get())
            horizontal_velocities = list(map(float, entries["horizontal_velocities"].get().split(',')))
            vertical_velocities = list(map(float, entries["vertical_velocities"].get().split(',')))

            vertical_error_input = entries["vertical_errors"].get().split(',')
            if not vertical_error_input[0]:  # Check if the first entry is empty
                vertical_error = [0.00001] * len(vertical_velocities)  # Default value array
            else:
                vertical_error = [float(x) if float(x) != 0 else 0.00001 for x in vertical_error_input]

            bird_data[bird_name] = {
                'parameters': {
                    'm': m,
                    'Sb': Sb,
                    'Sw': Sw,
                    'B': B
                },
                'velocities': {
                    'Horizontal': horizontal_velocities,
                    'Vertical': vertical_velocities,
                    'Vertical_error': vertical_error
                },
                'preferences': {
                    'save_fig': bool(save_fig_var.get()),
                    'save_csv': bool(save_csv_var.get()),
                    'save_excel': bool(save_excel_var.get())
                }
            }

            # Show save location
            if bird_data[bird_name]['preferences']['save_csv'] or bird_data[bird_name]['preferences']['save_fig']:
                messagebox.showinfo("Save Location", f"Data will be saved to: {folder_path}")

            more_birds = messagebox.askyesno("Continue to Advance Optimization", "Would you like to add more birds to optimize together in a combined calculation?")
            if not more_birds:
                root.quit()  # Close the tkinter window when 'No' is clicked
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter valid numeric values.")

    submit_button = tk.Button(root, text="Submit", command=submit_data)
    #submit_button.grid(row=len(labels), column=1, padx=pad_x, pady=pad_y, sticky=tk.EW)
    submit_button.grid(row=len(labels)+1, column=1, padx=pad_x, pady=pad_y, sticky=tk.EW)

    root.mainloop()

    return bird_data

objective = input_bird_data()


baseline_parameters = {'Cdb': 0.1, 'Cdpro': 0.014}
CDBmin = 0.1
CDBmax = 0.3

CDpromin = 0.003
CDpromax = 0.1

optimized_results = []
best_fitting_parameters_per_bird = {}

induced_drag_factor = 1.1



def wrapper_multiple(p, objective, baseline_parameters, debug=False, penalty_weight=0.01):
    loss_list = []
    for bird, objective_dict in objective.items():
        loss = wrapper(p, objective_dict=objective_dict, baseline_parameters=baseline_parameters, g=9.8, rho=1.01, debug=debug, penalty_weight=penalty_weight)
        loss_list.append(loss)
        if bird not in best_fitting_parameters_per_bird or loss < best_fitting_parameters_per_bird[bird]['loss']:
            best_fitting_parameters_per_bird[bird] = {
                'Cdb': CDBmin + p[0] * (CDBmax - CDBmin),
                'Cdpro': CDpromin + p[1] * (CDpromax - CDpromin),
                'loss': loss
            }
    overall_loss = np.mean(loss_list)
    return overall_loss

def plot_debug(p, loss, context):
    global objective
    for bird, objective_dict in objective.items():
        m = objective_dict['parameters']['m']
        Sb = objective_dict['parameters']['Sb']
        Sw = objective_dict['parameters']['Sw']
        B = objective_dict['parameters']['B']
        Cdb = CDBmin + p[0] * (CDBmax - CDBmin)
        Cdpro = CDpromin + p[1] * (CDpromax - CDpromin)

        new_results = f'{bird}, {Cdpro:.2g}, {Cdb:.2g}'
        optimized_results.append(new_results)

    print(p, loss, msg_dict[context])


def wrapper(p, objective_dict, baseline_parameters, g=9.8, rho=1.01, debug=False, penalty_weight=0.01):
    m = objective_dict['parameters']['m']
    Sb = objective_dict['parameters']['Sb']
    Sw = objective_dict['parameters']['Sw']
    B = objective_dict['parameters']['B']

    Cdb = CDBmin + p[0] * (CDBmax - CDBmin)
    Cdpro = CDpromin + p[1] * (CDpromax - CDpromin)

    delta = 1
    e = 1

    vz_array = []
    Vh_array = np.array(objective_dict["velocities"]['Horizontal']).astype(float)
    for i in Vh_array:
        vz_array.append(Vz_pennycuick(i, m=m, g=g, rho=rho, Sb=Sb, Sw=Sw, B=B, Cdpro=Cdpro, Cdb=Cdb, delta=delta, e=e))
    vz_array_objective = - np.array(objective_dict["velocities"]['Vertical']).astype(float)
    vz_array_err_objective = np.array(objective_dict["velocities"]['Vertical_error']).astype(float)
    vz_array = np.array(vz_array)

    # Calculate the primary loss (root-mean-square error)
    loss = np.sqrt(np.sum((vz_array_objective - vz_array) ** 2 / vz_array_err_objective) / np.sum(1 / vz_array_err_objective))

    # Apply a weighted penalty to prioritize Cdpro optimization
    # Here, we'll give Cdpro 3x the weight of Cdb
    penalty_weight_cdpro = penalty_weight * 3
    penalty_weight_cdb = penalty_weight

    # Compute soft penalties
    soft_penalty_cdpro = penalty_weight_cdpro * ((Cdpro - CDpromin) ** 2 + (Cdpro - CDpromax) ** 2) / ((CDpromax - CDpromin) ** 2)
    soft_penalty_cdb = penalty_weight_cdb * ((Cdb - CDBmin) ** 2 + (Cdb - CDBmax) ** 2) / ((CDBmax - CDBmin) ** 2)

    # Combine the loss with the penalties
    total_loss = loss + soft_penalty_cdpro + soft_penalty_cdb

    return total_loss

def wrapper_both(p):
    return wrapper_multiple([p[0], p[1]], objective, baseline_parameters)

def optimize_cdpro_only(objective, baseline_parameters, Cdb_fixed):
    def wrapper_cdpro_only(p_cdpro):
        p = [Cdb_fixed, p_cdpro[0]]
        return wrapper_multiple(p, objective, baseline_parameters)

    initial_guess = [(CDpromax + CDpromin) / 2]
    bounds = [(0, 1)]
    result = minimize(wrapper_cdpro_only, initial_guess, bounds=bounds)
    return result.x[0]

def optimize_cdb_only(objective, baseline_parameters, Cdpro_fixed):
    def wrapper_cdb_only(p_cdb):
        p = [p_cdb[0], Cdpro_fixed]  # Fix Cdpro, optimize only Cdb
        return wrapper_multiple(p, objective, baseline_parameters)

    initial_guess = [(CDBmax + CDBmin) / 2]
    bounds = [(0, 1)]
    result = minimize(wrapper_cdb_only, initial_guess, bounds=bounds)
    return result.x[0]  # Optimized Cdb

def full_optimization_process(objective, baseline_parameters):
    Cdb_fixed = (CDBmax + CDBmin) / 2

    # Stage 1: Optimize Cdpro
    optimized_Cdpro = optimize_cdpro_only(objective, baseline_parameters, Cdb_fixed)

    # Stage 2: Optimize Cdb
    optimized_Cdb = optimize_cdb_only(objective, baseline_parameters, optimized_Cdpro)

    return optimized_Cdb, optimized_Cdpro


result = minimize(wrapper_multiple, args=(objective, baseline_parameters), x0=[(CDBmax + CDBmin) / 2, 0.01],
                  method="CG")
optimized_Cdb, optimized_Cdpro = result.x
# Optimized Cdb
optimized_Cdb_final =CDBmin + optimized_Cdb * (CDBmax - CDBmin)
optimized_Cdpro_final =CDpromin + optimized_Cdpro * (CDpromax - CDpromin)
# Run the full optimization process
#optimized_Cdb, optimized_Cdpro = full_optimization_process(objective, baseline_parameters)

print(f"Optimized Cdb: {CDBmin + optimized_Cdb * (CDBmax - CDBmin)}")
print(f"Optimized Cdpro: {CDpromin + optimized_Cdpro  * (CDpromax - CDpromin)}")

my_bounds = [[0, 1], [0, 1]]
x0 = np.array([0.5, 0.5])
msg_dict = {0: "minimum detected in the annealing process.",
            1: "detection occurred in the local search process.",
            2: "detection done in the dual annealing process."}

penalty_weight = 0.01

best_parameters = {}
for bird in best_fitting_parameters_per_bird:
    best_parameters[bird] = best_fitting_parameters_per_bird[bird]

def closed_form_polar_curve(Vt, m, g, Sb, Sw, e, rho, delta, B, Cdpro, Cdb):
    AR = B ** 2 / Sw
    k = induced_drag_factor
    B_hat = beta_hat(Vt=Vt, m=m, g=g, rho=rho, Sw=Sw, B=B, k=k, Cdpro=Cdpro, delta=delta)
    if B_hat < 1:
        B0 = ((8 * k * m ** 2 * g ** 2) / (delta * np.pi * rho ** 2 * B ** 2 * Cdpro * Sw)) ** (1 / 3)
        Vz = 1 / (m * g) * (((2 * k * m ** 2 * g ** 2) / (np.pi * B0 ** 2 * B ** 2 * rho) * Vt ** (5 / 3)) + ((1 - delta) * rho * Sw * Cdpro / 2 * Vt ** 3) + (delta * rho * Sw * Cdpro / 2 * B0 * Vt ** (5 / 3)) + (rho * Sb * Cdb / 2 * Vt ** 3))
        return Vz
    else:
        Vz = 1 / (m * g) * (((2 * k * m ** 2 * g ** 2) / (np.pi * Vt ** (8 / 3) * B ** 2 * rho) * Vt ** (5 / 3)) + ((1 - delta) * rho * Sw * Cdpro / 2 * Vt ** 3) + (delta * rho * Sw * Cdpro / 2 * Vt ** (4 / 3) * Vt ** (5 / 3)) + (rho * Sb * Cdb / 2 * Vt ** 3))
        return Vz

def closed_form_derivative(Vt, m, g, Sb, Sw, e, rho, delta, B, Cdpro, Cdb):
    AR = B ** 2 / Sw
    k = induced_drag_factor
    B_hat = beta_hat(Vt=Vt, m=m, g=g, rho=rho, Sw=Sw, B=B, k=k, Cdpro=Cdpro, delta=delta)
    if B_hat < 1:
        B0 = ((8 * k * m ** 2 * g ** 2) / (delta * np.pi * rho ** 2 * B ** 2 * Cdpro * Sw)) ** (1 / 3)
        Vz_prime = 1 / (m * g) * (5 / 3 * 2 * k * m ** 2 * g ** 2 / (np.pi * B0 ** 2 * B ** 2 * rho) * Vt ** (2 / 3) + (3 * (1 - delta) * rho * Sw * Cdpro / 2 * Vt ** 2) + (5 / 3 * (delta * rho * Sw * Cdpro / 2) * B0 * Vt ** (2 / 3)) + (3 * rho * Sb * Cdb / 2 * Vt ** 2))
        return Vz_prime
    else:
        Vz_prime = 1 / (m * g) * ((-2 * k * m ** 2 * g ** 2) / (np.pi * B ** 2 * rho * Vt ** 2) + (3 * (1 - delta) * rho * Sw * Cdpro / 2 * Vt ** 2) + ((3 / 2 * delta * rho * Sw * Cdpro) * (Vt ** 2)) + (3 * rho * Sb * Cdb / 2 * Vt ** 2))
        return Vz_prime

def optimal_strategy(Vt, m, g, Sb, Sw, e, rho, delta, B, Cdpro, Cdb):
    climbrate = closed_form_polar_curve(Vt=Vt, m=m, g=g, Sb=Sb, Sw=Sw, e=e, rho=rho, delta=delta, B=B, Cdpro=Cdpro, Cdb=Cdb) - Vt * closed_form_derivative(Vt=Vt, m=m, g=g, Sb=Sb, Sw=Sw, e=e, rho=rho, delta=delta, B=B, Cdpro=Cdpro, Cdb=Cdb)
    return climbrate

def result(objective, best_parameters, g=9.8, rho=1.01):
    for bird, objective_dict in objective.items():
        m = objective_dict['parameters']['m']
        Sb = objective_dict['parameters']['Sb']
        Sw = objective_dict['parameters']['Sw']
        B = objective_dict['parameters']['B']
        Cdb = best_parameters[bird]['Cdb']
        Cdpro = best_parameters[bird]['Cdpro']
        loss = best_parameters[bird]['loss']
        delta = 1
        e = 1
        preferences = objective_dict["preferences"]


        Vh_array = np.array(objective_dict["velocities"]['Horizontal']).astype(float)
        vz_array = [Vz_pennycuick(i, m=m, g=g, rho=rho, Sb=Sb, Sw=Sw, B=B, Cdpro=Cdpro, Cdb=Cdb, delta=delta, e=e) for i in Vh_array]

        range_array = np.arange(5, 30.1, 0.1)
        standart_range = np.array(
            [Vz_pennycuick(i, m=m, g=g, rho=rho, Sb=Sb, Sw=Sw, B=B, Cdpro=Cdpro, Cdb=Cdb, delta=delta, e=e) for i in
             range_array])

        glide_polar = [
            closed_form_polar_curve(i, m=m, g=g, rho=rho, Sb=Sb, Sw=Sw, B=B, Cdpro=Cdpro, Cdb=Cdb, delta=delta, e=e) for
            i in
            range_array]

        derivative = [
            closed_form_derivative(i, m=m, g=g, rho=rho, Sb=Sb, Sw=Sw, B=B, Cdpro=Cdpro, Cdb=Cdb, delta=delta, e=e) for
            i in
            range_array]

        Vclimb = [optimal_strategy(i, m=m, g=g, rho=rho, Sb=Sb, Sw=Sw, B=B, Cdpro=Cdpro, Cdb=Cdb, delta=delta, e=e) for
                  i in range_array]

        opt_array = np.arange(0.01, 30.01, 0.01)

        Vclimb2 = [optimal_strategy(i, m=m, g=g, rho=rho, Sb=Sb, Sw=Sw, B=B, Cdpro=Cdpro, Cdb=Cdb, delta=delta, e=e) for
                   i in
                   opt_array]

        polar_from_zero = [
            closed_form_polar_curve(i, m=m, g=g, rho=rho, Sb=Sb, Sw=Sw, B=B, Cdpro=Cdpro, Cdb=Cdb, delta=delta, e=e) for
            i in
            opt_array]

        vz_array_objective = np.array(objective_dict["velocities"]['Vertical']).astype(float)
        vz_array_err_objective = np.array(objective_dict["velocities"]['Vertical_error']).astype(float)
        glide_polar = np.array(glide_polar) * -1
        derivative = np.array(derivative) * -1
        Vclimb = np.array(Vclimb) * -1
        standart_range = -standart_range
        Vclimb2 = np.array(Vclimb2) * -1

        polar_from_zero = np.array(polar_from_zero) * -1

        best_glide_speed_0 = opt_array[np.argmin(np.abs(Vclimb2 - 0))]
        best_glide_speed_1 = opt_array[np.argmin(np.abs(Vclimb2 - 1))]
        best_glide_speed_2 = opt_array[np.argmin(np.abs(Vclimb2 - 2))]
        best_glide_speed_3 = opt_array[np.argmin(np.abs(Vclimb2 - 3))]

        min_sink_x = np.where(polar_from_zero == max(polar_from_zero))
        min_sink_x = int(min_sink_x[0][0])
        speed_at_min_sink = opt_array[min_sink_x]
        min_sink_speed = closed_form_polar_curve(speed_at_min_sink, m=m, g=g, rho=rho, Sb=Sb, Sw=Sw, B=B, Cdpro=Cdpro,
                                                 Cdb=Cdb, delta=delta, e=e)
        min_sink_speed = -min_sink_speed

        best_glide_y = closed_form_polar_curve(best_glide_speed_0, m=m, g=g, rho=rho, Sb=Sb, Sw=Sw, B=B, Cdpro=Cdpro,
                                               Cdb=Cdb, delta=delta, e=e)
        best_glide_y = -best_glide_y

        best_glide_ratio = abs(best_glide_speed_0 / best_glide_y)

        linear_fit, _ = sc.optimize.curve_fit(f_linear, [0, 1, 2, 3],
                                              [best_glide_speed_0, best_glide_speed_1, best_glide_speed_2,
                                               best_glide_speed_3])
        a_linear, b_linear = linear_fit
        x_linear_fit = f_linear(np.arange(0, 3.1, 0.1), a_linear, b_linear)

        fig = plt.figure()
        plt.cla()
        plt.grid()

        plt.errorbar(Vh_array, vz_array_objective, vz_array_err_objective, fmt='o',label="observed gliding points")

        plt.plot(x_linear_fit, np.arange(0, 3.1, 0.1), c="red")


        plt.plot(range_array, glide_polar, c="red",label="estimated glide polar")

        plt.title(f"Glide Polar of {bird}", fontsize=20)
        plt.scatter(speed_at_min_sink, min_sink_speed, marker="o", c="red",label="minimum sink")
        plt.scatter(best_glide_speed_0, best_glide_y, marker="^", c="red", label= "maximum glide")

        plt.xlabel("Horizontal Speed (m/s)",fontsize=15)
        plt.ylabel("Vertical Speed (m/s)",fontsize=15)

        plt.legend()
        plt.ylim((-3, 0))
        plt.xlim((0, 20.5))
        fig.set_size_inches(8, 11)

        if preferences['save_fig']:
            save_path = os.path.join(folder_path, f"{bird}_Polar.png")

            fig.savefig(save_path, dpi=300)
            messagebox.showinfo("Save Successful", f"Successfully saved the figure to: {save_path}")

        plt.show(block=True)

        results = bird,m,B,Sw,Sb,best_glide_speed_0,best_glide_y, best_glide_ratio, speed_at_min_sink, min_sink_speed,Cdb, Cdpro,loss

        if preferences['save_csv']:
            save_csv = os.path.join(folder_path, f"{bird}_polar_results.csv")
            header = ["Input: Bird Name","Input: Mass(kg)","Input: Wingspan(m)","Input: Wing Area(m²)","Input: Body Frontal Area(m²)","Output: Speed at Best Glide (m/s)", "Output: Sink at Best Glide (m/s)", "Output: Best Glide Ratio", "Output: Speed at Min Sink (m/s)",
                      "Output: Min Sink Speed (m/s)","Output: Cdb", "Output: Cdpro", "Output: Loss"]

            with open(save_csv, "a+", newline="", encoding="utf-8-sig") as result_updater:
                csv_writer = csv.writer(result_updater)
                csv_writer.writerow(header)
                csv_writer.writerow(results)
                result_updater.close()
                messagebox.showinfo("Save Successful", f"Successfully saved CSV data to: {save_csv}")

        if preferences['save_excel']:
            save_xlsx = os.path.join(folder_path, f"{bird}_polar_results.xlsx")
            data_results = {
                "Bird Name": [results[0]],
                "Mass(kg)": [results[1]],
                "Wingspan(m)": [results[2]],
                "Wing Area(m²)": [results[3]],
                "Body Frontal Area(m²)": [results[4]],
                "Speed at Best Glide (m/s)": [results[5]],
                "Sink at Best Glide (m/s)": [results[6]],
                "Best Glide Ratio": [results[7]],
                "Speed at Min Sink (m/s)": [results[8]],
                "Min Sink Speed (m/s)": [results[9]],
                "Cdb": [results[10]],
                "Cdpro": [results[11]],
                "Loss": [results[12]]
            }

            results_df = pd.DataFrame(data_results)
            writer = pd.ExcelWriter(save_xlsx, engine='openpyxl')
            results_df.to_excel(writer, index=False, sheet_name='GlideTool',startrow=2)
            workbook = writer.book
            worksheet = writer.sheets['GlideTool']

            # Set the column widths
            column_widths = [15, 10, 12, 15, 20, 18, 18, 15, 18, 15, 10, 10, 10]
            col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
            for col_letter, width in zip(col_letters, column_widths):
                worksheet.column_dimensions[col_letter].width = width


            worksheet.merge_cells('A1:S1')  # Merges cells for input columns
            worksheet['A1'] = "#This is the GlideTool created as described in 'Data-driven optimization of Pennycuick’s Flight Tool for improved polar curves of thermal soaring bird species.'"


            worksheet.merge_cells('A2:E2')  # Merges cells for input columns
            worksheet.merge_cells('F2:M2')

            # Set the values for the merged header cells
            worksheet['A2'] = 'Input'
            worksheet['F2'] = 'Results'

            worksheet["A1"].alignment = Alignment(horizontal='left')

            for cell in ['A2','F2']:
                worksheet[cell].alignment = Alignment(horizontal='center')

            writer.close()
            messagebox.showinfo("Save Successful", f"Successfully saved Excel data to: {save_xlsx}")

result(objective, best_parameters)
